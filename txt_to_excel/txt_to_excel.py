# -*- coding: utf-8 -*-
#!/usr/bin/env python3

import sys
import argparse
import glob
from openpyxl import Workbook
import re
import time
import csv
import pkg_resources
import datetime

version = "1.2.1"

args = None
workbook = Workbook()
worksheet = workbook.active
date_data = []

def main():
    setup_arguments()

    setup_worksheet()
    
    if (sys.stdin.isatty()):
        # No input from stdin, read whole directory
        process_directory()
    else:
        process_stdin()

# Process files
def process_directory():
    files = glob.glob("**/*.txt", recursive=True) if args.use_recursion else glob.glob("./*.txt")
    process_files(files)

def process_stdin():
    raw_input = sys.stdin.readlines()
    input = []
    for line in raw_input:
        input.append(line.strip())

    process_files(input)

def process_files(files):
    print("Found {} files, getting to work.".format(len(files)))

    #read database of dates so we only do this once
    data_file = pkg_resources.resource_filename(__name__, 'data/date_data.csv')
    with open(data_file, "r") as dates_file:
        reader_stream = csv.DictReader(dates_file, delimiter=';')
        for row in reader_stream:
            date_data.append(row)

    for index, file in enumerate(files, start=1):
        progress = round(index / len(files) * 100)

        if not args.verbose: print ("Progress: [{:<20}] {}%".format("–"*(round(progress/5)), progress), end="\r")
        process_file(index, file)

    #Setup sorting
    worksheet.auto_filter.ref = "A1:E{}".format(worksheet.max_row)
    worksheet.auto_filter.add_sort_condition("E1:E{}".format(worksheet.max_row))
    #print("A1:{}{}".format(chr(65+worksheet.max_column), worksheet.max_row))

    workbook.save(args.filename)

    print("✅ Successfully imported {} files into {}".format(len(files), args.filename))

def process_file(index, file):
    if args.verbose: print("📝 Reading file Nr. {}, {:<40} ".format(index, file), end="", flush=True)
    
    with open(file, "r") as stream:
        data = stream.read()
    
    #global worksheet
    data_to = search_for_pattern(data, "TO: (.+?)\n")
    data_from = search_for_pattern(data, "FROM: (.+?)\n")
    data_subject = search_for_pattern(data, "SUBJECT: (.+?)\n")
    data_message = search_for_pattern(data, "\n\n.*$", 0)
    data_date = lookup_date(file)

    worksheet.append([data_to, data_from, data_subject, data_message, data_date])

    if args.verbose: print("✅")

def search_for_pattern(data, pattern, group=1):    
    try:
        return re.search(pattern, data, re.DOTALL).group(group)
    except AttributeError:
        return ""

def lookup_date(file):
    file_id = re.sub(r"\D", "", file)

    for line in date_data:
        if line['Ticket-ID'] == file_id:
            date_time = "{} {}".format(line['Datum'], line['Uhrzeit'])
            return datetime.datetime.strptime(date_time, "%d.%m.%y %H:%M:%S")   
        
    return("")


# Setup
def setup_arguments():
    usage_examples = '''Example usages:
    Make all files in this directory into an Excel file called "abc.xlsx":
        txt-to-excel --filename abc.xlsx

    Make all files in the current directory that include #Seenotrettung into an Excel file using grep:
        grep -rl "#Seenotrettung" . | txt-to-excel -f seenotrettung.xlsx
    '''

    parser = argparse.ArgumentParser(description="Process txt into an excel worksheet. If no input is specified, the current directory is searched for txt files.", epilog=usage_examples, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('-r', dest="use_recursion", action="store_true", help="Search recursively, i.e. consider subdirectories. By default, the script only looks in the current directory.")
    parser.add_argument('-f', "--filename", nargs="?", dest="filename", default="output.xlsx", help="The filename for the output excel file. Default is 'output.xlsx'")
    parser.add_argument('-v', "--verbose", dest="verbose", action="store_true", help="Print the progress and other useful info.")
    parser.add_argument('--version', action='version', version="Program version: {}".format(version))

    global args
    args = parser.parse_args()

def setup_worksheet():
    # Set titles for our Columns
    headline_style = "Headline 2"
    column_width_small = 25
    column_width_medium = 50
    column_width_large = 80

    worksheet["A1"] = "To"
    worksheet["A1"].style = headline_style
    worksheet.column_dimensions["A"].width = column_width_small

    worksheet["B1"] = "From"
    worksheet["B1"].style = headline_style
    worksheet.column_dimensions["B"].width = column_width_medium

    worksheet["C1"] = "Subject"
    worksheet["C1"].style = headline_style
    worksheet.column_dimensions["C"].width = column_width_small

    worksheet["D1"] = "Text"
    worksheet["D1"].style = headline_style
    worksheet.column_dimensions["D"].width = column_width_large

    worksheet["E1"] = "Timestamp"
    worksheet["E1"].style = headline_style
    worksheet.column_dimensions["E"].width = column_width_small